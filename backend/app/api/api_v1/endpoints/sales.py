from typing import List, Optional
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
import pandas as pd
from sqlalchemy.orm import Session
import os
import shutil
from pathlib import Path
from app.api.deps import get_db, get_current_user
from app.core.file_storage import file_storage
import pandas as pd
import json
from app.models.user import User
from app.models.sale import Sale, SaleItem, SaleStatus
from app.models.product import Product
from app.models.inventory import Inventory
from app.models.notification import Notification, NotificationType
from app.schemas.sale import (
    SaleCreate,
    SaleUpdate,
    Sale as SaleSchema,
    SaleList
)
from decimal import Decimal
import uuid

router = APIRouter()

@router.get("/", response_model=SaleList)
def get_sales(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=10000),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[List[str]] = Query(None),
    brand_name: Optional[List[str]] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """판매 목록 조회"""
    from sqlalchemy.orm import selectinload
    from app.models.brand import Brand

    # Sale을 조회하면서 관련 데이터를 미리 로드
    query = db.query(Sale)\
        .join(User, Sale.seller_id == User.id)\
        .options(
            selectinload(Sale.items).selectinload(SaleItem.product).selectinload(Product.brand),
            selectinload(Sale.seller)
        )

    # 필터링
    if start_date:
        query = query.filter(Sale.sale_date >= start_date)
    if end_date:
        query = query.filter(Sale.sale_date <= end_date)
    if status:
        # 다중 선택 지원
        status_enums = [SaleStatus(s) for s in status]
        query = query.filter(Sale.status.in_(status_enums))
    if brand_name:
        # 다중 선택 지원 - subquery 사용하여 DISTINCT/ORDER BY 충돌 방지
        brand_sale_ids = db.query(Sale.id)\
            .join(Sale.items)\
            .join(SaleItem.product)\
            .join(Product.brand)\
            .filter(Brand.name.in_(brand_name))\
            .distinct()\
            .subquery()
        query = query.filter(Sale.id.in_(brand_sale_ids))

    # seller 권한은 자신의 판매만 조회
    if current_user.role.value == "seller":
        query = query.filter(Sale.seller_id == current_user.id)

    total = query.count()

    # 판매일자 최신순 정렬
    query = query.order_by(Sale.sale_date.desc())
    sales = query.offset(skip).limit(limit).all()

    # seller_name 필드 추가 및 product 정보 동적 할당
    for sale in sales:
        if sale.seller:
            sale.seller_name = sale.seller.full_name

        # SaleItem의 product 정보를 동적으로 할당
        for item in sale.items:
            if item.product:
                # 동적으로 product_code 속성 추가 (스키마에만 존재)
                setattr(item, 'product_code', item.product.product_code)
                if not item.product_name:
                    item.product_name = item.product.product_name

                # brand_name 추가
                if item.product.brand:
                    setattr(item, 'brand_name', item.product.brand.name)

                # product_image_url이 없으면 product 정보로부터 생성
                if not item.product_image_url and item.product.brand and item.product.product_code:
                    brand_name = item.product.brand.name
                    product_code = item.product.product_code
                    item.product_image_url = f"/uploads/products/{brand_name}/{product_code}.png"

    return SaleList(total=total, items=sales)

@router.get("/{sale_id}", response_model=SaleSchema)
def get_sale(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """판매 상세 조회"""
    from sqlalchemy.orm import selectinload
    from app.models.brand import Brand

    sale = db.query(Sale)\
        .options(
            selectinload(Sale.items).selectinload(SaleItem.product).selectinload(Product.brand),
            selectinload(Sale.seller)
        )\
        .filter(Sale.id == sale_id).first()

    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    # seller_name 필드 추가
    if sale.seller:
        sale.seller_name = sale.seller.full_name

    # SaleItem의 product 정보를 동적으로 할당
    for item in sale.items:
        if item.product:
            # 동적으로 product_code 속성 추가 (스키마에만 존재)
            setattr(item, 'product_code', item.product.product_code)
            if not item.product_name:
                item.product_name = item.product.product_name

            # brand_name 추가
            if item.product.brand:
                setattr(item, 'brand_name', item.product.brand.name)

            # category 추가
            if item.product.category:
                setattr(item, 'category', item.product.category)

            # product_image_url이 없으면 product 정보로부터 생성
            if not item.product_image_url and item.product.brand and item.product.product_code:
                brand_name = item.product.brand.name
                product_code = item.product.product_code
                item.product_image_url = f"/uploads/products/{brand_name}/{product_code}.png"

    return sale

@router.post("/", response_model=SaleSchema)
def create_sale(
    sale_data: SaleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """판매 등록"""
    # seller나 admin만 판매 등록 가능
    if current_user.role.value not in ["seller", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # 판매번호 생성 (S + 날짜 + 순번)
    from datetime import datetime
    today_str = datetime.now().strftime("%Y%m%d")

    # 오늘 등록된 판매 중 가장 큰 번호 조회
    last_sale = db.query(Sale).filter(
        Sale.sale_number.like(f"S{today_str}-%")
    ).order_by(Sale.sale_number.desc()).first()

    if last_sale and last_sale.sale_number:
        # 마지막 번호에서 순번 추출
        last_number = int(last_sale.sale_number.split('-')[-1])
        next_number = last_number + 1
    else:
        next_number = 1

    sale_number = f"S{today_str}-{next_number:04d}"

    # 판매 생성
    sale = Sale(
        id=uuid.uuid4(),
        sale_number=sale_number,
        sale_date=sale_data.sale_date,
        seller_id=current_user.id,
        customer_name=sale_data.customer_name,
        customer_contact=sale_data.customer_contact,
        notes=sale_data.notes,
        status=SaleStatus.pending,
        total_seller_amount=Decimal(0),
        total_company_amount=Decimal(0),
        total_seller_margin=Decimal(0)
    )

    db.add(sale)

    # 판매 아이템 생성 및 총액 계산
    total_seller_amount = Decimal(0)

    for item_data in sale_data.items:
        # 상품 확인
        product = db.query(Product).filter(Product.id == item_data.product_id).first()
        if not product:
            raise HTTPException(status_code=400, detail=f"Product {item_data.product_id} not found")

        # 재고 확인 (사이즈별로 체크)
        inventory = db.query(Inventory).filter(
            Inventory.product_id == item_data.product_id,
            Inventory.size == item_data.size
        ).first()

        if not inventory:
            raise HTTPException(
                status_code=400,
                detail=f"Inventory not found for product {product.product_name} size {item_data.size}"
            )

        if inventory.available_quantity < item_data.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient inventory for product {product.product_name} size {item_data.size}. Available: {inventory.available_quantity}, Requested: {item_data.quantity}"
            )

        item = SaleItem(
            id=uuid.uuid4(),
            sale_id=sale.id,
            product_id=item_data.product_id,
            product_name=item_data.product_name or product.product_name,
            size=item_data.size,
            quantity=item_data.quantity,
            seller_sale_price_original=item_data.seller_sale_price_original,
            seller_sale_currency=item_data.seller_sale_currency,
            seller_sale_price_krw=item_data.seller_sale_price_krw,
            product_image_url=item_data.product_image_url
        )

        db.add(item)
        total_seller_amount += item_data.seller_sale_price_krw * item_data.quantity

        # 재고 차감 및 알림 생성
        if inventory:
            previous_quantity = inventory.quantity
            new_quantity = previous_quantity - item_data.quantity
            inventory.quantity = new_quantity

            # 알림 생성 로직
            # 이미지 URL 생성 (item_data에 있으면 사용, 없으면 브랜드/상품코드로 생성)
            image_url = item_data.product_image_url
            if not image_url and product.brand:
                image_url = f"/uploads/products/{product.brand.name}/{product.product_code}.png"

            # 품절 알림: 이전 재고가 1개 이상이었다가 0개가 됨
            if previous_quantity >= 1 and new_quantity == 0:
                notification = Notification(
                    id=str(uuid.uuid4()),
                    type=NotificationType.STOCK_OUT,
                    title="품절 알림",
                    message=f"{product.product_name} ({product.product_code}) - {item_data.size} 사이즈가 품절되었습니다.",
                    product_id=str(product.id),
                    product_name=product.product_name,
                    product_code=product.product_code,
                    product_image_url=image_url,
                    size=item_data.size,
                    previous_quantity=str(previous_quantity),
                    current_quantity=str(new_quantity)
                )
                db.add(notification)

            # 재고 부족 알림: 이전 재고가 6개 이상이었다가 5개 이하로 떨어짐
            elif previous_quantity >= 6 and new_quantity <= 5 and new_quantity > 0:
                notification = Notification(
                    id=str(uuid.uuid4()),
                    type=NotificationType.STOCK_LOW,
                    title="재고 부족 알림",
                    message=f"{product.product_name} ({product.product_code}) - {item_data.size} 사이즈의 재고가 {new_quantity}개로 부족합니다.",
                    product_id=str(product.id),
                    product_name=product.product_name,
                    product_code=product.product_code,
                    product_image_url=image_url,
                    size=item_data.size,
                    previous_quantity=str(previous_quantity),
                    current_quantity=str(new_quantity)
                )
                db.add(notification)

    # 총액 업데이트
    sale.total_seller_amount = total_seller_amount

    db.commit()
    db.refresh(sale)

    return sale

@router.put("/{sale_id}", response_model=SaleSchema)
def update_sale(
    sale_id: str,
    sale_update: SaleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """판매 정보 수정"""
    import logging
    logger = logging.getLogger(__name__)

    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    # 업데이트
    update_data = sale_update.dict(exclude_unset=True)
    logger.info(f"Updating sale {sale_id} with data: {update_data}")

    for field, value in update_data.items():
        setattr(sale, field, value)

    # 회사 판매가가 업데이트된 경우 판매자 마진 자동 계산 및 상태 업데이트
    if 'total_company_amount' in update_data and sale.total_seller_amount:
        sale.total_seller_margin = sale.total_seller_amount - sale.total_company_amount
        logger.info(f"Calculated margin: {sale.total_seller_margin} = {sale.total_seller_amount} - {sale.total_company_amount}")

        # 회사 판매가가 입력되면 상태를 completed로 변경
        if sale.total_company_amount and sale.total_company_amount > 0:
            sale.status = SaleStatus.completed
            logger.info(f"Status updated to COMPLETED")
        else:
            sale.status = SaleStatus.pending
            logger.info(f"Status updated to PENDING")

    db.commit()
    db.refresh(sale)

    logger.info(f"Sale {sale_id} updated successfully. Company amount: {sale.total_company_amount}")

    return sale

@router.delete("/{sale_id}")
def delete_sale(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """판매 삭제"""
    # admin만 삭제 가능
    if current_user.role.value != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 재고 원복
    items = db.query(SaleItem).filter(SaleItem.sale_id == sale_id).all()
    for item in items:
        inventory = db.query(Inventory).filter(
            Inventory.product_id == item.product_id
        ).first()
        if inventory:
            inventory.quantity += item.quantity

    db.delete(sale)
    db.commit()

    return {"message": "Sale deleted successfully"}

@router.patch("/items/{item_id}")
def update_sale_item(
    item_id: str,
    company_sale_price: Optional[Decimal] = None,
    seller_margin: Optional[Decimal] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """판매 아이템 업데이트 (관리자용)"""
    # admin만 가능
    if current_user.role.value != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    item = db.query(SaleItem).filter(SaleItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Sale item not found")

    # 업데이트
    if company_sale_price is not None:
        item.company_sale_price = company_sale_price
    if seller_margin is not None:
        item.seller_margin = seller_margin

    db.commit()
    db.refresh(item)

    return {"message": "Sale item updated successfully"}

@router.post("/{sale_id}/upload-transaction-statement")
async def upload_transaction_statement(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    file: UploadFile = File(...)
):
    """거래명세서 업로드"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Uploading transaction statement for sale {sale_id}")
    logger.info(f"File: {file.filename}, Content-Type: {file.content_type}")
    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    # 파일 저장
    try:
        custom_name = f"sale_{sale_id}_statement"
        relative_path = await file_storage.save_file(
            file=file,
            file_type='sale_transaction_statement',
            custom_name=custom_name
        )

        # DB 업데이트
        sale.transaction_statement_url = relative_path
        db.commit()

        return {
            "message": "Transaction statement uploaded successfully",
            "file_path": relative_path,
            "url": file_storage.get_file_url(relative_path, "/api/v1/files")
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to upload file: {str(e)}")

@router.post("/{sale_id}/upload-tax-invoice")
async def upload_tax_invoice(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    file: UploadFile = File(...)
):
    """세금계산서 업로드"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Uploading tax invoice for sale {sale_id}")
    logger.info(f"File: {file.filename}, Content-Type: {file.content_type}")
    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    # 파일 저장
    try:
        custom_name = f"sale_{sale_id}_tax"
        relative_path = await file_storage.save_file(
            file=file,
            file_type='sale_tax_invoice',
            custom_name=custom_name
        )

        # DB 업데이트
        sale.tax_invoice_url = relative_path
        db.commit()

        return {
            "message": "Tax invoice uploaded successfully",
            "file_path": relative_path,
            "url": file_storage.get_file_url(relative_path, "/api/v1/files")
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to upload file: {str(e)}")

@router.get("/{sale_id}/transaction-statement")
def download_transaction_statement(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """거래명세서 다운로드"""
    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    if not sale.transaction_statement_url:
        raise HTTPException(status_code=404, detail="Transaction statement not found")

    # 파일 경로 가져오기
    file_path = file_storage.get_full_path(sale.transaction_statement_url)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        filename=file_path.name,
        media_type='application/octet-stream'
    )

@router.get("/{sale_id}/transaction-statement-preview")
def preview_transaction_statement(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """거래명세서 미리보기 (엑셀 내용 반환)"""
    from openpyxl import load_workbook
    import json

    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    if not sale.transaction_statement_url:
        raise HTTPException(status_code=404, detail="Transaction statement not found")

    # 파일 경로 가져오기
    file_path = file_storage.get_full_path(sale.transaction_statement_url)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    try:
        # 파일 확장자에 따라 다른 방법으로 읽기
        file_extension = file_path.suffix.lower()

        if file_extension == '.csv':
            # CSV 파일은 기존 방식으로 읽기
            df = pd.read_csv(file_path, encoding='utf-8-sig')
            df = df.fillna('')
            preview_df = df.head(100)

            return {
                "type": "simple",
                "columns": preview_df.columns.tolist(),
                "data": preview_df.values.tolist(),
                "total_rows": len(df),
                "file_name": file_path.name
            }

        elif file_extension in ['.xlsx', '.xls']:
            # Excel 파일은 openpyxl로 상세하게 읽기
            workbook = load_workbook(filename=str(file_path), data_only=True)
            sheet = workbook.active

            # 셀 데이터와 스타일 정보 수집
            sheet_data = []
            merged_cells = []

            # 병합된 셀 정보 수집
            for merge_range in sheet.merged_cells.ranges:
                merged_cells.append({
                    "min_row": merge_range.min_row,
                    "max_row": merge_range.max_row,
                    "min_col": merge_range.min_col,
                    "max_col": merge_range.max_col
                })

            # 최대 100행까지만 읽기
            max_row = min(sheet.max_row, 100)
            max_col = sheet.max_column

            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_info = {
                        "value": cell.value if cell.value is not None else "",
                        "row": row_idx,
                        "col": col_idx
                    }

                    # 스타일 정보 추가
                    if cell.font:
                        cell_info["bold"] = cell.font.bold
                        cell_info["italic"] = cell.font.italic
                        if cell.font.color and cell.font.color.rgb:
                            cell_info["color"] = f"#{cell.font.color.rgb[2:]}" if isinstance(cell.font.color.rgb, str) else None

                    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                        bg_color = cell.fill.start_color.rgb
                        if isinstance(bg_color, str) and bg_color != "00000000":
                            cell_info["backgroundColor"] = f"#{bg_color[2:]}"

                    if cell.alignment:
                        cell_info["align"] = cell.alignment.horizontal
                        cell_info["valign"] = cell.alignment.vertical

                    if cell.border:
                        borders = {}
                        if cell.border.top and cell.border.top.style:
                            borders["top"] = True
                        if cell.border.bottom and cell.border.bottom.style:
                            borders["bottom"] = True
                        if cell.border.left and cell.border.left.style:
                            borders["left"] = True
                        if cell.border.right and cell.border.right.style:
                            borders["right"] = True
                        if borders:
                            cell_info["borders"] = borders

                    row_data.append(cell_info)
                sheet_data.append(row_data)

            workbook.close()

            return {
                "type": "styled",
                "data": sheet_data,
                "merged_cells": merged_cells,
                "total_rows": sheet.max_row,
                "total_cols": max_col,
                "file_name": file_path.name
            }
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {str(e)}")

@router.post("/{sale_id}/return")
def process_return(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """반품 처리 - 재고 원복 후 상태 변경"""
    import logging
    logger = logging.getLogger(__name__)

    # seller 또는 admin만 반품 처리 가능
    if current_user.role.value not in ["seller", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크 (seller는 자신의 판매만)
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    # 이미 반품 처리된 경우
    if sale.status == SaleStatus.returned:
        raise HTTPException(status_code=400, detail="이미 반품 처리된 판매건입니다.")

    # 재고 원복
    items = db.query(SaleItem).filter(SaleItem.sale_id == sale_id).all()
    for item in items:
        inventory = db.query(Inventory).filter(
            Inventory.product_id == item.product_id,
            Inventory.size == item.size
        ).first()

        if inventory:
            logger.info(f"Restoring inventory for product {item.product_id}, size {item.size}: {inventory.quantity} + {item.quantity}")
            inventory.quantity += item.quantity
        else:
            # 재고 레코드가 없으면 새로 생성
            logger.info(f"Creating new inventory for product {item.product_id}, size {item.size}: {item.quantity}")
            new_inventory = Inventory(
                id=uuid.uuid4(),
                product_id=item.product_id,
                size=item.size,
                quantity=item.quantity
            )
            db.add(new_inventory)

    # 상태 변경
    sale.status = SaleStatus.returned

    db.commit()
    db.refresh(sale)

    logger.info(f"Sale {sale_id} returned successfully. Items restored to inventory.")

    return {"message": "반품 처리가 완료되었습니다. 재고가 원복되었습니다."}


@router.post("/{sale_id}/cancel-return")
def cancel_return(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """반품 취소 처리 - 재고 다시 차감 후 상태 복원"""
    import logging
    logger = logging.getLogger(__name__)

    # admin만 반품 취소 가능
    if current_user.role.value != "admin":
        raise HTTPException(status_code=403, detail="관리자만 반품 취소가 가능합니다.")

    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 반품 상태가 아닌 경우
    if sale.status != SaleStatus.returned:
        raise HTTPException(status_code=400, detail="반품 상태인 판매건만 취소할 수 있습니다.")

    # 재고 다시 차감
    items = db.query(SaleItem).filter(SaleItem.sale_id == sale_id).all()
    for item in items:
        inventory = db.query(Inventory).filter(
            Inventory.product_id == item.product_id,
            Inventory.size == item.size
        ).first()

        if inventory:
            if inventory.quantity < item.quantity:
                raise HTTPException(
                    status_code=400,
                    detail=f"재고가 부족하여 반품 취소가 불가합니다. (상품: {item.product_name}, 사이즈: {item.size})"
                )
            logger.info(f"Deducting inventory for product {item.product_id}, size {item.size}: {inventory.quantity} - {item.quantity}")
            inventory.quantity -= item.quantity
        else:
            raise HTTPException(
                status_code=400,
                detail=f"재고 정보가 없어 반품 취소가 불가합니다. (상품: {item.product_name}, 사이즈: {item.size})"
            )

    # 상태 복원 (회사 판매가가 있으면 completed, 없으면 pending)
    if sale.total_company_amount and sale.total_company_amount > 0:
        sale.status = SaleStatus.completed
    else:
        sale.status = SaleStatus.pending

    db.commit()
    db.refresh(sale)

    logger.info(f"Sale {sale_id} return cancelled. Items deducted from inventory.")

    return {"message": "반품 취소가 완료되었습니다. 재고가 다시 차감되었습니다."}


@router.get("/{sale_id}/tax-invoice")
def download_tax_invoice(
    sale_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """세금계산서 다운로드"""
    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # 권한 체크
    if current_user.role.value == "seller" and str(sale.seller_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Not authorized")

    if not sale.tax_invoice_url:
        raise HTTPException(status_code=404, detail="Tax invoice not found")

    # 파일 경로 가져오기
    file_path = file_storage.get_full_path(sale.tax_invoice_url)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        filename=file_path.name,
        media_type='application/octet-stream'
    )